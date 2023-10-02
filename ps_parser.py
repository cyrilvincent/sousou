from typing import List, Optional
from dataclasses import dataclass
import openpyxl.reader.excel
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
import argparse
import shutil
import os
import sys
import csv



@dataclass
class DSN:

    id: str
    value: str

@dataclass
class Salarie:

    row: int
    etablissement: str
    matricule: str
    nom: str
    prenom: str
    statut: str
    profession: str
    libelle: str
    nature: str
    salaire_base: str #'010'
    montant: float

class DSNParser:

    def __init__(self, path: str):
        self.path = path
        self.row = 0
        self.dsns: List[DSN] = []
        self.etab = ""

    def parse(self):
        print(f"Chargement de {self.path}")
        self.parse_etab()
        with open(self.path) as f:
            reader = csv.reader(f)
            for row in reader:
                self.row += 1
                id = row[0]
                value = row[1][1:-1]
                dsn = DSN(id, value)
                self.dsns.append(dsn)
        print(f"{len(self.dsns)} lignes chargées")

    def parse_etab(self):
        try:
            index_ = self.path.rindex("-")
            index_dot = self.path.rindex(".")
            self.etab = self.path[index_ + 1:index_dot]
            print(f"Etablissement {self.etab}")
        except:
            print("Mauvais format ddu nom de fichier, etablissement impossible à déduire")

class XLWriter:

    def __init__(self, template_path: str, out_path: str):
        self.template_path = template_path
        self.out_path = out_path
        self.wb: Optional[Workbook] = None
        self.sheet: Optional[Worksheet] = None
        self.row = 3

    def create_and_load(self):
        if not os.path.isfile(self.template_path):
            print(f"Erreur: {self.template_path} n'existe pas")
            sys.exit(1)
        if not os.path.isfile(self.out_path):
            print(f"Création de {self.out_path}")
            shutil.copy2(self.template_path, self.out_path)
        self.wb = openpyxl.reader.excel.load_workbook(self.out_path)

    def create_and_load_sheet(self, name: str):
        if name not in self.wb.sheetnames:
            sheet = self.wb["template"]
            self.wb.copy_worksheet(sheet)
        self.sheet = self.wb[self.wb.sheetnames[-1]]
        self.sheet.title = name

    def write_salarie(self, s: Salarie):
        self.sheet.cell(self.row, 1, s.etablissement)
        self.sheet.cell(self.row, 2, s.matricule)
        self.sheet.cell(self.row, 3, s.nom)
        self.sheet.cell(self.row, 4, s.prenom)
        self.sheet.cell(self.row, 5, s.statut)
        self.sheet.cell(self.row, 6, s.profession)
        self.sheet.cell(self.row, 7, s.libelle)
        self.sheet.cell(self.row, 8, s.nature)
        self.sheet.cell(self.row, 9, s.salaire_base)
        self.sheet.cell(self.row, 10, s.montant)

    def write(self, salaries: List[Salarie]):
        print(f"Ecriture dans la feuille {self.sheet.title}")
        for s in salaries:
            self.write_salarie(s)
            self.row += 1

    def save(self):
        print(f"Sauvegarde de {self.out_path}")
        try:
            self.wb.save(self.out_path)
        except:
            input(f"Merci de fermer {self.out_path} et d'appuyer sur Entrée")
            self.wb.save(self.out_path)

    def remove_template(self):
        self.wb = openpyxl.reader.excel.load_workbook(self.out_path)
        self.wb.remove(self.wb["template"])
        self.save()



class DSNService:

    def __init__(self, dsn_path, xl_template_path, xl_path):
        self.dsn_parser = DSNParser(dsn_path)
        self.xl_writer = XLWriter(xl_template_path, xl_path)
        self.row = 0
        self.nb = 0
        self.salaries: List[Salarie] = []

    def start(self):
        self.xl_writer.create_and_load()
        self.dsn_parser.parse()
        while self.row < len(self.dsn_parser.dsns):
            dsn = self.dsn_parser.dsns[self.row]
            if dsn.id == "S21.G00.30.002":
                self.parse_salarie(dsn)
            self.row += 1
        print(f"Nb salarié: {self.nb}")
        self.xl_writer.create_and_load_sheet(self.dsn_parser.etab)
        self.xl_writer.write(self.salaries)
        self.xl_writer.save()

    def parse_salarie(self, dsn: DSN):
        print(f"Salarié ligne {self.row + 1} : {dsn.value}")
        self.nb += 1
        nom = dsn.value
        row = self.row
        self.row += 1
        nom = prenom = matricule = statut = profession = libelle = nature = salaire = ""
        montant = 0
        is_montant_010 = False
        while self.row < len(self.dsn_parser.dsns):
            dsn = self.dsn_parser.dsns[self.row]
            if dsn.id == "S21.G00.30.004":
                prenom = dsn.value
            elif dsn.id == "S21.G00.30.019":
                matricule = dsn.value
            elif dsn.id == "S21.G00.40.003":
                statut = dsn.value
            elif dsn.id == "S21.G00.40.004":
                profession = dsn.value
            elif dsn.id == "S21.G00.40.006":
                libelle = dsn.value
            elif dsn.id == "S21.G00.40.007":
                nature = dsn.value
            elif dsn.id == "S21.G00.51.011":
                if dsn.value == "010":
                    salaire = dsn.value
                    is_montant_010 = True
                else:
                    is_montant_010 = False
            elif dsn.id == "S21.G00.51.013" and is_montant_010:
                montant = float(dsn.value)
            elif dsn.id == "S21.G00.30.002":
                self.row -= 1
                break
            self.row += 1
        salarie = Salarie(row, self.dsn_parser.etab, matricule, nom, prenom, statut, profession, libelle
                          , nature, salaire, montant)
        self.salaries.append(salarie)

class DSNDirectoryService:

    def __init__(self, dsn_directory_path: str, xl_template_path: str):
        self.dsn_directory_path = dsn_directory_path
        self.xl_template_path = xl_template_path
        self.nb = 0
        self.nb_salarie = 0

    def guess_xl_name(self):
        xl_name = self.dsn_directory_path.replace("\\", "/")
        if xl_name.__contains__("/"):
            xl_name = xl_name.split("/")[-1]
        xl_name += ".xlsx"
        return xl_name

    def backup_and_create(self, xl_name: str):
        if os.path.isfile(xl_name):
            print(f"Sauvegarde de {xl_name} dans {xl_name}.bak")
            shutil.copy2(xl_name, xl_name+".bak")
            try:
                os.remove(xl_name)
            except:
                input(f"Merci de fermer le fichier {xl_name} et d'appuyer sur entrée")
                os.remove(xl_name)

    def start(self):
        print(f"Lecture du répertoire {self.dsn_directory_path}")
        if not os.path.isdir(self.dsn_directory_path):
            print(f"{self.dsn_directory_path} doit être un répertoire")
            sys.exit(2)
        xl_name = self.guess_xl_name()
        self.backup_and_create(xl_name)
        l = os.listdir(self.dsn_directory_path)
        l.sort()
        for f in l:
            if f.upper().endswith(".TXT") or f.upper().endswith(".DSN"):
                self.nb += 1
                s = DSNService( self.dsn_directory_path + "/" + f, self.xl_template_path, xl_name)
                s.start()
                self.nb_salarie += len(s.salaries)
        s = XLWriter(self.xl_template_path, xl_name)
        s.remove_template()
        print(f"Lecture de {self.nb} fichier(s) DSN et création de {self.nb_salarie} salarié(s)")





if __name__ == '__main__':
    print("DSN Import")
    print("==========")
    print()
    parser = argparse.ArgumentParser(description="DSN Import")
    parser.add_argument("dsn_path", help="DSN Directory Path")
    args = parser.parse_args()
    s = DSNDirectoryService(args.dsn_path, "template.xlsx")
    s.start()

